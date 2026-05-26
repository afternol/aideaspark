# DBから全アイデアをExcelに出力するスクリプト
import psycopg2
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

import os
try:
    from dotenv import load_dotenv; load_dotenv()
except ImportError:
    pass
DATABASE_URL = os.environ.get("DATABASE_URL", "")
OUTPUT_PATH = r"C:\Users\after\AI副業_YouTube\AideaSpark_アイデア200件.xlsx"

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()
cur.execute("""
    SELECT
        number, id, slug, "serviceName", "oneLiner", concept, target, problem,
        product, "revenueModel", competitors, "competitiveEdge",
        tags, category, "targetIndustry", "targetCustomer",
        "investmentScale", difficulty, scores, "scoreComments",
        "trendKeywords", patterns, "whyNow", "noveltyNote", "strengthNote",
        "patternRationale", "publishedAt", views, bookmarks,
        "createdAt", "updatedAt"
    FROM "Idea"
    ORDER BY number ASC
""")
rows = cur.fetchall()
print(f"取得件数: {len(rows)}")

col_names = [
    "number", "id", "slug", "serviceName", "oneLiner", "concept", "target", "problem",
    "product", "revenueModel", "competitors", "competitiveEdge",
    "tags", "category", "targetIndustry", "targetCustomer",
    "investmentScale", "difficulty",
    "score_novelty", "score_marketSize", "score_profitability", "score_growth", "score_feasibility", "score_moat",
    "comment_novelty", "comment_marketSize", "comment_profitability", "comment_growth", "comment_feasibility", "comment_moat",
    "trendKeywords", "patterns", "whyNow", "noveltyNote", "strengthNote",
    "patternRationale", "publishedAt", "views", "bookmarks",
    "createdAt", "updatedAt"
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "アイデア一覧"

# ヘッダースタイル
header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for ci, name in enumerate(col_names, 1):
    cell = ws.cell(row=1, column=ci, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 30

# データ行
alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
normal_font = Font(name="Meiryo UI", size=9)

for ri, row in enumerate(rows, 2):
    (number, id_, slug, serviceName, oneLiner, concept, target, problem,
     product, revenueModel, competitors, competitiveEdge,
     tags, category, targetIndustry, targetCustomer,
     investmentScale, difficulty, scores, scoreComments,
     trendKeywords, patterns, whyNow, noveltyNote, strengthNote,
     patternRationale, publishedAt, views, bookmarks,
     createdAt, updatedAt) = row

    # scoresをパース
    if isinstance(scores, str):
        scores = json.loads(scores)
    if isinstance(scoreComments, str):
        scoreComments = json.loads(scoreComments)
    if isinstance(tags, str):
        tags = json.loads(tags)
    if isinstance(trendKeywords, str):
        trendKeywords = json.loads(trendKeywords)
    if isinstance(patterns, str):
        patterns = json.loads(patterns)

    def jl(v):
        if isinstance(v, list):
            return ", ".join(str(x) for x in v)
        return str(v) if v else ""

    values = [
        number, id_, slug, serviceName, oneLiner, concept, target, problem,
        product, revenueModel, competitors, competitiveEdge,
        jl(tags), category, targetIndustry, targetCustomer,
        investmentScale, difficulty,
        scores.get("novelty", ""), scores.get("marketSize", ""),
        scores.get("profitability", ""), scores.get("growth", ""),
        scores.get("feasibility", ""), scores.get("moat", ""),
        scoreComments.get("novelty", ""), scoreComments.get("marketSize", ""),
        scoreComments.get("profitability", ""), scoreComments.get("growth", ""),
        scoreComments.get("feasibility", ""), scoreComments.get("moat", ""),
        jl(trendKeywords), jl(patterns), whyNow, noveltyNote, strengthNote,
        patternRationale, publishedAt, views, bookmarks,
        str(createdAt), str(updatedAt)
    ]

    fill = alt_fill if ri % 2 == 0 else None
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font = normal_font
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        cell.border = border
        if fill:
            cell.fill = fill

# 列幅の設定
col_widths = {
    1: 7,    # number
    2: 28,   # id
    3: 20,   # slug
    4: 20,   # serviceName
    5: 30,   # oneLiner
    6: 50,   # concept
    7: 30,   # target
    8: 40,   # problem
    9: 50,   # product
    10: 40,  # revenueModel
    11: 30,  # competitors
    12: 40,  # competitiveEdge
    13: 20,  # tags
    14: 15,  # category
    15: 15,  # targetIndustry
    16: 15,  # targetCustomer
    17: 12,  # investmentScale
    18: 8,   # difficulty
    19: 8, 20: 8, 21: 8, 22: 8, 23: 8, 24: 8,   # scores
    25: 35, 26: 35, 27: 35, 28: 35, 29: 35, 30: 35,  # scoreComments
    31: 20,  # trendKeywords
    32: 15,  # patterns
    33: 50,  # whyNow
    34: 50,  # noveltyNote
    35: 50,  # strengthNote
    36: 40,  # patternRationale
    37: 12,  # publishedAt
    38: 7,   # views
    39: 7,   # bookmarks
    40: 20,  # createdAt
    41: 20,  # updatedAt
}
for ci, w in col_widths.items():
    ws.column_dimensions[get_column_letter(ci)].width = w

# ウィンドウ枠を固定（1行目・1列目）
ws.freeze_panes = "B2"

# フィルター設定
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_PATH)
print("OK: " + OUTPUT_PATH)

cur.close()
conn.close()
